# -*- coding: utf-8 -*-
"""
航班数据爬虫系统 - Excel输出处理模块

本模块提供航班数据的Excel输出功能：
- 数据格式化和验证
- Excel文件创建和写入
- 多航段数据整理
- 统计信息生成
- 文件保存和管理
"""

import os
import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .logger import get_logger, log_exception, log_flight_process
from .config_manager import get_config_manager


class OutputHandler:
    """
    输出处理器
    
    负责将提取的航班数据输出到Excel文件
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化输出处理器
        
        Args:
            config: 配置字典
        """
        self.logger = get_logger()
        
        if config is None:
            config_manager = get_config_manager()
            self.output_config = config_manager.get_output_config()
        else:
            self.output_config = config.get('output', {})
        
        # 输出目录
        self.output_dir = self.output_config.get('directory', 'output')
        if not os.path.isabs(self.output_dir):
            self.output_dir = os.path.abspath(self.output_dir)
        
        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Excel列定义
        self.excel_columns = {
            'flight_number': '航班号',
            'departure_date': '出发日期',
            'segment_index': '航段序号',
            'departure_airport': '出发机场',
            'arrival_airport': '到达机场',
            'scheduled_departure': '计划起飞时间',
            'scheduled_arrival': '计划到达时间',
            'actual_departure': '实际起飞时间',
            'actual_arrival': '实际到达时间',
            'flight_status': '航班状态',
            'created_time': '数据获取时间'
        }
    
    def save_flight_data(self, flight_data: List[Dict[str, Any]], 
                        filename: Optional[str] = None) -> Dict[str, Any]:
        """
        保存航班数据到Excel文件
        
        Args:
            flight_data: 航班数据列表
            filename: 输出文件名（可选）
            
        Returns:
            保存结果字典
        """
        try:
            if not flight_data:
                return {
                    'success': False,
                    'message': '没有数据需要保存',
                    'file_path': None,
                    'statistics': None
                }
            
            # 生成文件名
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'flight_data_{timestamp}.xlsx'
            
            file_path = os.path.join(self.output_dir, filename)
            
            if self.logger:
                self.logger.info(f"开始保存航班数据到文件: {file_path}")
            
            # 数据预处理
            processed_data = self._preprocess_data(flight_data)
            
            # 创建Excel文件
            result = self._create_excel_file(processed_data, file_path)
            
            if result['success']:
                # 生成统计信息
                statistics = self._generate_statistics(processed_data)
                result['statistics'] = statistics
                
                if self.logger:
                    self.logger.info(f"航班数据保存成功: {file_path}")
                    self.logger.info(f"统计信息: 总航段 {statistics['total_segments']}，有效航段 {statistics['valid_segments']}")
            
            return result
            
        except Exception as e:
            error_msg = f"保存航班数据失败: {str(e)}"
            if self.logger:
                log_exception('output_handler', 'save_flight_data', e, 
                            {'filename': filename, 'data_count': len(flight_data) if flight_data else 0})
            
            return {
                'success': False,
                'message': error_msg,
                'file_path': None,
                'statistics': None
            }
    
    def _preprocess_data(self, flight_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        预处理航班数据
        
        Args:
            flight_data: 原始航班数据
            
        Returns:
            处理后的数据
        """
        processed_data = []
        
        for data in flight_data:
            processed_item = {}
            
            # 复制所有字段
            for key, value in data.items():
                processed_item[key] = value
            
            # 数据清理和格式化
            processed_item = self._clean_data_item(processed_item)
            
            processed_data.append(processed_item)
        
        # 按航班号和航段序号排序
        processed_data.sort(key=lambda x: (x.get('flight_number', ''), x.get('segment_index', 0)))
        
        return processed_data
    
    def _clean_data_item(self, data_item: Dict[str, Any]) -> Dict[str, Any]:
        """
        清理单个数据项
        
        Args:
            data_item: 数据项
            
        Returns:
            清理后的数据项
        """
        cleaned_item = data_item.copy()
        
        # 处理空值
        for key, value in cleaned_item.items():
            if value is None:
                cleaned_item[key] = ''
            elif isinstance(value, str):
                # 清理字符串
                cleaned_item[key] = value.strip()
        
        # 确保必需字段存在
        required_fields = ['flight_number', 'departure_date', 'segment_index']
        for field in required_fields:
            if field not in cleaned_item:
                cleaned_item[field] = ''
        
        # 格式化航段序号
        try:
            segment_index = cleaned_item.get('segment_index', 0)
            if isinstance(segment_index, str):
                cleaned_item['segment_index'] = int(segment_index) if segment_index.isdigit() else 0
        except (ValueError, TypeError):
            cleaned_item['segment_index'] = 0
        
        return cleaned_item
    
    def _create_excel_file(self, data: List[Dict[str, Any]], file_path: str) -> Dict[str, Any]:
        """
        创建Excel文件
        
        Args:
            data: 处理后的数据
            file_path: 文件路径
            
        Returns:
            创建结果
        """
        try:
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 重新排列列顺序
            ordered_columns = []
            for col_key in self.excel_columns.keys():
                if col_key in df.columns:
                    ordered_columns.append(col_key)
            
            # 添加其他列
            for col in df.columns:
                if col not in ordered_columns:
                    ordered_columns.append(col)
            
            df = df[ordered_columns]
            
            # 重命名列
            column_mapping = {}
            for col in df.columns:
                if col in self.excel_columns:
                    column_mapping[col] = self.excel_columns[col]
                else:
                    column_mapping[col] = col
            
            df = df.rename(columns=column_mapping)
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "航班数据"
            
            # 写入数据
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 应用样式
            self._apply_excel_styles(ws, len(df) + 1)
            
            # 保存文件
            wb.save(file_path)
            
            return {
                'success': True,
                'message': '文件创建成功',
                'file_path': file_path,
                'row_count': len(df)
            }
            
        except Exception as e:
            if self.logger:
                log_exception('output_handler', '_create_excel_file', e, {'file_path': file_path})
            
            return {
                'success': False,
                'message': f'Excel文件创建失败: {str(e)}',
                'file_path': file_path,
                'row_count': 0
            }
    
    def _apply_excel_styles(self, worksheet, row_count: int):
        """
        应用Excel样式
        
        Args:
            worksheet: 工作表对象
            row_count: 行数
        """
        try:
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            data_alignment = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 应用标题行样式
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 应用数据行样式
            for row in range(2, row_count + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            if self.logger:
                log_exception('output_handler', '_apply_excel_styles', e)
    
    def _generate_statistics(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        生成统计信息
        
        Args:
            data: 航班数据
            
        Returns:
            统计信息字典
        """
        try:
            if not data:
                return {
                    'total_segments': 0,
                    'valid_segments': 0,
                    'flights_count': 0,
                    'success_rate': 0.0,
                    'field_statistics': {}
                }
            
            total_segments = len(data)
            valid_segments = 0
            flights = set()
            
            # 字段统计
            field_stats = {}
            check_fields = ['departure_airport', 'arrival_airport', 'scheduled_departure', 
                          'scheduled_arrival', 'actual_departure', 'actual_arrival', 'flight_status']
            
            for field in check_fields:
                field_stats[field] = {
                    'total': 0,
                    'success': 0,
                    'success_rate': 0.0
                }
            
            # 遍历数据统计
            for item in data:
                # 统计航班数
                flight_key = f"{item.get('flight_number', '')}_{item.get('departure_date', '')}"
                flights.add(flight_key)
                
                # 判断是否为有效航段
                is_valid = bool(
                    item.get('flight_number') and 
                    item.get('departure_airport') and 
                    item.get('arrival_airport')
                )
                
                if is_valid:
                    valid_segments += 1
                
                # 统计各字段
                for field in check_fields:
                    field_stats[field]['total'] += 1
                    value = item.get(field, '')
                    
                    if value and value not in ['识别失败', '元素未找到', '提取异常', '图片保存失败', '']:
                        field_stats[field]['success'] += 1
            
            # 计算成功率
            success_rate = (valid_segments / total_segments * 100) if total_segments > 0 else 0
            
            # 计算字段成功率
            for field in field_stats:
                total = field_stats[field]['total']
                success = field_stats[field]['success']
                field_stats[field]['success_rate'] = (success / total * 100) if total > 0 else 0
            
            return {
                'total_segments': total_segments,
                'valid_segments': valid_segments,
                'flights_count': len(flights),
                'success_rate': round(success_rate, 2),
                'field_statistics': {k: {
                    'total': v['total'],
                    'success': v['success'],
                    'success_rate': round(v['success_rate'], 2)
                } for k, v in field_stats.items()}
            }
            
        except Exception as e:
            if self.logger:
                log_exception('output_handler', '_generate_statistics', e)
            
            return {
                'total_segments': len(data) if data else 0,
                'valid_segments': 0,
                'flights_count': 0,
                'success_rate': 0.0,
                'field_statistics': {}
            }
    
    def append_to_existing_file(self, flight_data: List[Dict[str, Any]], 
                               existing_file: str) -> Dict[str, Any]:
        """
        追加数据到现有Excel文件
        
        Args:
            flight_data: 新的航班数据
            existing_file: 现有文件路径
            
        Returns:
            操作结果
        """
        try:
            if not os.path.exists(existing_file):
                return self.save_flight_data(flight_data, os.path.basename(existing_file))
            
            # 读取现有数据
            existing_df = pd.read_excel(existing_file)
            
            # 预处理新数据
            processed_data = self._preprocess_data(flight_data)
            new_df = pd.DataFrame(processed_data)
            
            # 重命名列以匹配现有文件
            column_mapping = {}
            for col in new_df.columns:
                if col in self.excel_columns:
                    column_mapping[col] = self.excel_columns[col]
            new_df = new_df.rename(columns=column_mapping)
            
            # 合并数据
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            # 保存合并后的数据
            combined_df.to_excel(existing_file, index=False)
            
            if self.logger:
                self.logger.info(f"数据追加成功: {existing_file}，新增 {len(flight_data)} 条记录")
            
            return {
                'success': True,
                'message': f'成功追加 {len(flight_data)} 条记录',
                'file_path': existing_file,
                'total_rows': len(combined_df)
            }
            
        except Exception as e:
            error_msg = f"追加数据失败: {str(e)}"
            if self.logger:
                log_exception('output_handler', 'append_to_existing_file', e, 
                            {'existing_file': existing_file, 'new_data_count': len(flight_data)})
            
            return {
                'success': False,
                'message': error_msg,
                'file_path': existing_file,
                'total_rows': 0
            }
    
    def create_summary_report(self, statistics: Dict[str, Any], 
                             output_file: Optional[str] = None) -> Dict[str, Any]:
        """
        创建汇总报告
        
        Args:
            statistics: 统计信息
            output_file: 输出文件名
            
        Returns:
            创建结果
        """
        try:
            if not output_file:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = f'flight_summary_{timestamp}.xlsx'
            
            file_path = os.path.join(self.output_dir, output_file)
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "数据汇总"
            
            # 写入汇总信息
            ws.append(['统计项目', '数值'])
            ws.append(['总航段数', statistics.get('total_segments', 0)])
            ws.append(['有效航段数', statistics.get('valid_segments', 0)])
            ws.append(['航班数量', statistics.get('flights_count', 0)])
            ws.append(['成功率(%)', statistics.get('success_rate', 0)])
            ws.append([])
            
            # 写入字段统计
            ws.append(['字段名称', '总数', '成功数', '成功率(%)'])
            field_stats = statistics.get('field_statistics', {})
            for field, stats in field_stats.items():
                field_name = self.excel_columns.get(field, field)
                ws.append([
                    field_name,
                    stats.get('total', 0),
                    stats.get('success', 0),
                    stats.get('success_rate', 0)
                ])
            
            # 应用样式
            self._apply_excel_styles(ws, ws.max_row)
            
            # 保存文件
            wb.save(file_path)
            
            if self.logger:
                self.logger.info(f"汇总报告创建成功: {file_path}")
            
            return {
                'success': True,
                'message': '汇总报告创建成功',
                'file_path': file_path
            }
            
        except Exception as e:
            error_msg = f"创建汇总报告失败: {str(e)}"
            if self.logger:
                log_exception('output_handler', 'create_summary_report', e, {'output_file': output_file})
            
            return {
                'success': False,
                'message': error_msg,
                'file_path': None
            }


# 便捷函数
def save_flight_data(flight_data: List[Dict[str, Any]], filename: Optional[str] = None) -> Dict[str, Any]:
    """
    保存航班数据的便捷函数
    
    Args:
        flight_data: 航班数据列表
        filename: 输出文件名
        
    Returns:
        保存结果
    """
    handler = OutputHandler()
    return handler.save_flight_data(flight_data, filename)


def create_summary_report(statistics: Dict[str, Any], output_file: Optional[str] = None) -> Dict[str, Any]:
    """
    创建汇总报告的便捷函数
    
    Args:
        statistics: 统计信息
        output_file: 输出文件名
        
    Returns:
        创建结果
    """
    handler = OutputHandler()
    return handler.create_summary_report(statistics, output_file)