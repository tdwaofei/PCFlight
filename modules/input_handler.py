# -*- coding: utf-8 -*-
"""
航班数据爬虫系统 - Excel输入处理模块

本模块提供Excel文件的读取和航班信息提取功能：
- Excel文件格式验证
- 航班号和日期信息提取
- 数据格式验证和清理
- 错误处理和日志记录
"""

import os
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import re

from .logger import get_logger, log_exception


class InputHandler:
    """
    Excel输入处理器
    
    负责读取和处理输入的Excel文件，提取航班信息
    """
    
    def __init__(self):
        """
        初始化输入处理器
        """
        self.logger = get_logger()
        self.supported_extensions = ['.xlsx', '.xls']
        # 航班号格式：前2位必须是字母或数字，后面3-4位必须是数字
        # 但不能是纯数字（如12345）或只有1位前缀（如M5100）
        # 例如：MU5100, G54381, 3U8888, CA1234等
        self.flight_number_pattern = re.compile(r'^[A-Z0-9]{2}\d{3,4}$')
        self.date_formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
            '%Y%m%d'
        ]
    
    def read_flight_data(self, file_path: str) -> List[Dict[str, str]]:
        """
        读取输入Excel文件中的航班信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            航班信息列表，每个元素包含flight_number和departure_date
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不支持或数据格式错误
        """
        try:
            # 验证文件存在性和格式
            self._validate_file(file_path)
            
            if self.logger:
                self.logger.info(f"开始读取Excel文件: {file_path}")
            
            # 读取Excel文件
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # 提取航班数据
            flight_data = self._extract_flight_data(worksheet)
            
            # 验证和清理数据
            validated_data = self._validate_and_clean_data(flight_data)
            
            if self.logger:
                self.logger.info(f"成功读取 {len(validated_data)} 条航班信息")
            
            workbook.close()
            return validated_data
            
        except Exception as e:
            if self.logger:
                log_exception('input_handler', 'read_flight_data', e, 
                            {'file_path': file_path})
            raise
    
    def _validate_file(self, file_path: str) -> None:
        """
        验证文件的存在性和格式
        
        Args:
            file_path: 文件路径
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不支持
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in self.supported_extensions:
            raise ValueError(f"不支持的文件格式: {file_ext}，支持的格式: {self.supported_extensions}")
        
        # 检查文件是否可读
        try:
            with open(file_path, 'rb') as f:
                f.read(1)
        except PermissionError:
            raise PermissionError(f"没有权限读取文件: {file_path}")
    
    def _extract_flight_data(self, worksheet) -> List[Dict[str, Any]]:
        """
        从工作表中提取航班数据
        
        Args:
            worksheet: openpyxl工作表对象
            
        Returns:
            原始航班数据列表
        """
        flight_data = []
        
        # 检测数据开始行（跳过标题行）
        start_row = self._detect_data_start_row(worksheet)
        
        if self.logger:
            self.logger.info(f"数据开始行: {start_row}")
        
        # 读取数据行
        for row_num in range(start_row, worksheet.max_row + 1):
            # 读取A列（航班号）和B列（出发日期）
            flight_number_cell = worksheet.cell(row=row_num, column=1)
            departure_date_cell = worksheet.cell(row=row_num, column=2)
            
            flight_number = flight_number_cell.value
            departure_date = departure_date_cell.value
            
            # 跳过空行
            if not flight_number and not departure_date:
                continue
            
            flight_data.append({
                'flight_number': flight_number,
                'departure_date': departure_date,
                'row_index': row_num,
                'original_flight_number': flight_number,
                'original_departure_date': departure_date
            })
        
        return flight_data
    
    def _detect_data_start_row(self, worksheet) -> int:
        """
        检测数据开始行（自动跳过标题行）
        
        Args:
            worksheet: openpyxl工作表对象
            
        Returns:
            数据开始行号（1-based）
        """
        # 检查前几行，寻找可能的标题行
        for row_num in range(1, min(6, worksheet.max_row + 1)):
            cell_a = worksheet.cell(row=row_num, column=1).value
            cell_b = worksheet.cell(row=row_num, column=2).value
            
            # 如果第一列包含"航班"或"flight"等关键词，认为是标题行
            if cell_a and isinstance(cell_a, str):
                if any(keyword in cell_a.lower() for keyword in ['航班', 'flight', '班次']):
                    return row_num + 1
            
            # 如果第二列包含"日期"或"date"等关键词，认为是标题行
            if cell_b and isinstance(cell_b, str):
                if any(keyword in cell_b.lower() for keyword in ['日期', 'date', '时间']):
                    return row_num + 1
        
        # 默认从第2行开始（假设第1行是标题）
        return 2
    
    def _validate_and_clean_data(self, raw_data: List[Dict[str, Any]]) -> List[Dict[str, str]]:
        """
        验证和清理航班数据
        
        Args:
            raw_data: 原始数据列表
            
        Returns:
            验证和清理后的数据列表
        """
        validated_data = []
        errors = []
        
        for i, item in enumerate(raw_data):
            try:
                # 清理和验证航班号
                flight_number = self._clean_and_validate_flight_number(
                    item['flight_number'], item['row_index']
                )
                
                # 清理和验证出发日期
                departure_date = self._clean_and_validate_date(
                    item['departure_date'], item['row_index']
                )
                
                if flight_number and departure_date:
                    validated_data.append({
                        'flight_number': flight_number,
                        'departure_date': departure_date,
                        'row_index': item['row_index']
                    })
                
            except Exception as e:
                error_msg = f"第{item['row_index']}行数据验证失败: {str(e)}"
                errors.append(error_msg)
                if self.logger:
                    self.logger.warning(error_msg)
        
        # 记录验证结果
        if self.logger:
            self.logger.info(f"数据验证完成: 总行数={len(raw_data)}, 有效数据={len(validated_data)}, 错误数据={len(errors)}")
            
            if errors:
                self.logger.warning(f"发现 {len(errors)} 个数据错误:")
                for error in errors[:10]:  # 只显示前10个错误
                    self.logger.warning(f"  - {error}")
                if len(errors) > 10:
                    self.logger.warning(f"  - ... 还有 {len(errors) - 10} 个错误")
        
        if not validated_data:
            raise ValueError("没有找到有效的航班数据")
        
        return validated_data
    
    def _clean_and_validate_flight_number(self, flight_number: Any, row_index: int) -> Optional[str]:
        """
        清理和验证航班号
        
        Args:
            flight_number: 原始航班号
            row_index: 行索引
            
        Returns:
            清理后的航班号，如果无效则返回None
        """
        if not flight_number:
            return None
        
        # 转换为字符串并清理基本空白字符
        flight_str = str(flight_number).strip()
        
        # 检查是否为空
        if not flight_str:
            return None
        
        # 检查是否包含不允许的字符（允许字母、数字，大小写均可）
        if not re.match(r'^[A-Za-z0-9]+$', flight_str):
            raise ValueError(f"航班号格式不正确: {flight_number} (包含无效字符，只允许字母和数字)")
        
        # 转换为大写进行后续验证
        flight_str = flight_str.upper()
        
        # 验证航班号格式：2位字母或数字 + 3-4位数字
        if not self.flight_number_pattern.match(flight_str):
            raise ValueError(f"航班号格式不正确: {flight_number} (应为2位字母或数字+3-4个数字，如MU5100、G54381、3U8888)")
        
        return flight_str
    
    def _clean_and_validate_date(self, departure_date: Any, row_index: int) -> Optional[str]:
        """
        清理和验证出发日期
        
        Args:
            departure_date: 原始出发日期
            row_index: 行索引
            
        Returns:
            格式化后的日期字符串（YYYY-MM-DD），如果无效则返回None
        """
        if not departure_date:
            return None
        
        # 如果是datetime对象，直接格式化
        if isinstance(departure_date, datetime):
            return departure_date.strftime('%Y-%m-%d')
        
        # 转换为字符串并清理
        date_str = str(departure_date).strip()
        
        # 尝试解析不同的日期格式
        for date_format in self.date_formats:
            try:
                parsed_date = datetime.strptime(date_str, date_format)
                
                # 验证日期合理性（不能是过去太久或未来太远的日期）
                current_date = datetime.now()
                days_diff = (parsed_date - current_date).days
                
                if days_diff < -365:  # 不能是一年前的日期
                    raise ValueError(f"日期过于久远: {date_str}")
                if days_diff > 365:  # 不能是一年后的日期
                    raise ValueError(f"日期过于遥远: {date_str}")
                
                return parsed_date.strftime('%Y-%m-%d')
                
            except ValueError:
                continue
        
        # 如果所有格式都无法解析
        raise ValueError(f"无法解析日期格式: {departure_date} (支持格式: {', '.join(self.date_formats)})")
    
    def create_sample_input_file(self, file_path: str, sample_data: Optional[List[Dict[str, str]]] = None) -> None:
        """
        创建示例输入文件
        
        Args:
            file_path: 输出文件路径
            sample_data: 示例数据，如果为None则使用默认示例
        """
        if sample_data is None:
            sample_data = [
                {'flight_number': 'MU5100', 'departure_date': '2025-09-17'},
                {'flight_number': 'CA1234', 'departure_date': '2025-09-18'},
                {'flight_number': 'CZ3456', 'departure_date': '2025-09-19'}
            ]
        
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "航班信息"
            
            # 写入标题行
            worksheet['A1'] = '航班号'
            worksheet['B1'] = '出发日期'
            
            # 写入数据行
            for i, data in enumerate(sample_data, start=2):
                worksheet[f'A{i}'] = data['flight_number']
                worksheet[f'B{i}'] = data['departure_date']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 15
            
            workbook.save(file_path)
            
            if self.logger:
                self.logger.info(f"示例输入文件创建成功: {file_path}")
                
        except Exception as e:
            if self.logger:
                log_exception('input_handler', 'create_sample_input_file', e, 
                            {'file_path': file_path})
            raise
    
    def validate_input_file_format(self, file_path: str) -> Tuple[bool, List[str]]:
        """
        验证输入文件格式是否正确
        
        Args:
            file_path: 文件路径
            
        Returns:
            (是否有效, 错误信息列表)
        """
        errors = []
        
        try:
            # 基本文件验证
            self._validate_file(file_path)
            
            # 读取文件并检查格式
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # 检查是否有数据
            if worksheet.max_row < 2:
                errors.append("文件中没有数据行")
            
            # 检查前几行数据格式
            start_row = self._detect_data_start_row(worksheet)
            sample_rows = min(5, worksheet.max_row - start_row + 1)
            
            for row_num in range(start_row, start_row + sample_rows):
                flight_number = worksheet.cell(row=row_num, column=1).value
                departure_date = worksheet.cell(row=row_num, column=2).value
                
                if not flight_number and not departure_date:
                    continue
                
                # 验证航班号格式
                try:
                    self._clean_and_validate_flight_number(flight_number, row_num)
                except Exception as e:
                    errors.append(f"第{row_num}行航班号格式错误: {str(e)}")
                
                # 验证日期格式
                try:
                    self._clean_and_validate_date(departure_date, row_num)
                except Exception as e:
                    errors.append(f"第{row_num}行日期格式错误: {str(e)}")
            
            workbook.close()
            
        except Exception as e:
            errors.append(f"文件读取错误: {str(e)}")
        
        return len(errors) == 0, errors
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        获取文件信息
        
        Args:
            file_path: 文件路径
            
        Returns:
            文件信息字典
        """
        try:
            file_stat = os.stat(file_path)
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            start_row = self._detect_data_start_row(worksheet)
            data_rows = max(0, worksheet.max_row - start_row + 1)
            
            info = {
                'file_path': file_path,
                'file_size': file_stat.st_size,
                'file_size_mb': round(file_stat.st_size / 1024 / 1024, 2),
                'modified_time': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'total_rows': worksheet.max_row,
                'data_rows': data_rows,
                'start_row': start_row,
                'worksheet_name': worksheet.title
            }
            
            workbook.close()
            return info
            
        except Exception as e:
            if self.logger:
                log_exception('input_handler', 'get_file_info', e, {'file_path': file_path})
            return {'error': str(e)}


# 便捷函数
def read_flight_data(file_path: str) -> List[Dict[str, str]]:
    """
    读取航班数据的便捷函数
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        航班信息列表
    """
    handler = InputHandler()
    return handler.read_flight_data(file_path)


def create_sample_input_file(file_path: str) -> None:
    """
    创建示例输入文件的便捷函数
    
    Args:
        file_path: 输出文件路径
    """
    handler = InputHandler()
    handler.create_sample_input_file(file_path)


def validate_input_file(file_path: str) -> Tuple[bool, List[str]]:
    """
    验证输入文件的便捷函数
    
    Args:
        file_path: 文件路径
        
    Returns:
        (是否有效, 错误信息列表)
    """
    handler = InputHandler()
    return handler.validate_input_file_format(file_path)